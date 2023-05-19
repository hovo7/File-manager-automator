from openpyxl import Workbook, load_workbook
# to get the letter of the columns
from openpyxl.utils import get_column_letter


#1) # create the excel sheet and comment out the code after executing it once to recreate a new file in every execution
# wb = Workbook()
# ws = wb.active
# ws.tile = "Data"

# ws.append(['A1','B1','C1','D1'])
# ws.append(['A2','B2','C2','D2'])
# ws.append(['A3','B3','C3','D3'])
# ws.append(['END'])

# wb.save('Test_WB.xlsx')

#2)
wb = load_workbook('Test_WB.xlsx')
ws = wb.active

for row in range(1,11):
    for col in range(1, 5):
        char = get_column_letter(col)
        location = ws[char + str(row)]   
        ws[char + str(row)] = char + str(row)
        

# merging and unmerging can be done too
# merging: ws.merge_cells("A1:D1")
# unmerging: ws.unmerge_cells("A1:D1")

# inserting and deleting rows are possible
# inserting: ws.insert_row(psotion)
# deleting: ws.delete_row(psotion)

# inserting and deleting columns are possible
# inserting: ws.insert_cols(psotion)
# deleting: ws.delete_cols(psotion)

# to move range of rows and cols
# ws.move_range(C1:D11, row = push up or down , cols= push left or right)

wb.save('Test_WB.xlsx')